"""
- Copy the schema listing made by elec department
- Parse it and export to Postgres DB
"""
import shutil
import pandas as pd
import openpyxl
import os
import urllib
from db_utils.base import Session


# Source file path
source_file = r"\\aludra\base_be_ELEC\base CLIENTS\Listing Schémas.xlsm"
# Destination directory
destination_directory = "./temp"

# Path to the copied Excel file
copied_file_path = "./temp/Listing Schémas.xlsm"



# Copy the file to the destination directory
shutil.copy(source_file, destination_directory)
print(f"File copied successfully to {destination_directory}")

# Sheet name to read
sheet_name = 'AF'

# Read the 'AF' sheet from the Excel file into a pandas DataFrame
df = pd.read_excel(copied_file_path, sheet_name=sheet_name)
df['lien'] = ''
wb = openpyxl.load_workbook(copied_file_path)
# sheets = wb.sheetnames
ws = wb[sheet_name]
for i, row in df.iterrows():
    n_row = i + 2
    link = None
    try:
        raw_path = '\\\\aludra\\base_be_ELEC\\base CLIENTS\\' + ws.cell(row=n_row, column=3).hyperlink.target
        decoded_path = urllib.parse.unquote(raw_path)
        fixed_path = os.path.normpath(decoded_path)
        df.loc[i, 'lien'] = fixed_path
    except:
        pass
df['code_equip'] = df['N° AF'] + '-' + df['LETTRAGE']
# df.drop(columns=['N° AF', 'LETTRAGE'], inplace=True)

# Keep only non empty code_equip
df = df.loc[df['code_equip'].isna() == False]

# Rename columns and select required columns
how_to_rename = {
    'N° AF': 'Num_AF',
    'LETTRAGE': 'Lettrage',
    'Nom Client': 'Client',
    'DATE': 'Date',
    'TENSION + FREQUENCE': 'Tension_Freq',
    'NEUTRE': 'Neutre',
    'TYPE': 'Type',
    'CARACTERISTIQUES': 'Caract',
    'SPECIFICITES': 'Specif',
    'N° SCHEMAS': 'Num_Schemas',
    'NBR MS=': 'Nb_MS',
    'MS1= ': 'MS1',
    'MS2=': 'MS2',
    'MS3=': 'MS3',
    'MS4=': 'MS4',
    'NBR ME=': 'Nb_ME',
    'ME1=': 'ME1',
    'ME2=': 'ME2',
    'ME3=': 'ME3',
    'ME4=': 'ME4',
    'DEM MOT=': 'Dem_Mot',
    'CHAUF=': 'Chauff',
    'REGUL=': 'Regul',
    'PILOT. RES ELEC=': 'Pilot_Res_Elec',
    'PUIS ELEC=': 'Puis_Elec',
    'HUMIDIF=': 'Humidif',
    'FROID=': 'Froid',
    'PAF= ': 'PAF',
    'RECY=': 'Recy',
    'EXT SOLV=': 'Ext_Solv',
    'PRES=': 'Pres',
    'DIAPH=': 'Diaph',
    'API= ': 'API',
    'IHM = ': 'IHM',
    'VERSION': 'Version',
    'DESSINATEUR': 'Dessinateur',
    'Nom de Programme Automate': 'Prog_Auto',
    'Nom de Programme Afficheur': 'Prog_Affich',
    'LANGUE': 'Langue',
    'ARMOIRE': 'Armoire',
    'Nb Cabine Identique': 'Nb_Cab_Ident',
    'lien': 'Lien',
    'code_equip': 'Code_Equip'
}


df = df.rename(columns=how_to_rename)
df = df.loc[:, list(how_to_rename.values())]
print(f"Dataframe ready to be exported to Postgres ({len(df)} rows)")
with Session() as session:
    df.to_sql('schema_special_plus', con=session.connection(), if_exists='replace', index=False, schema='dimensional_ax')
    session.commit()
print("Dataframe exported to Postgres")